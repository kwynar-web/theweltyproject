#!/usr/bin/env python3
"""
build_records.py — wire the Record Images/ backlog onto the family tree.

For every file in  Record Images/  this script:
  1. converts it to a web JPG (PDFs via pdftoppm) at two sizes ->
        site/records/full/<slug>.jpg   (max 1600px)
        site/records/thumb/<slug>.jpg  (max 360px)
  2. reads the curated per-file MAPPING below (which person[s] on the tree
     the record belongs to, a caption, a confidence flag)
  3. pulls the repository + external URL from  Record Images/Capture Manifest.xlsx
  4. writes  site/records/records.json   -> { PersonID: [ {slug,caption,repo,url,logids,confidence}, ... ] }
     plus an "_evidence" bucket for records that document a decoy / negative
     sweep / separate cluster and MUST NOT be pinned to a direct-line person.
  5. mirrors the whole map into a readable "Tree Image Map" sheet in the
     Capture Manifest workbook so the curation is durable and reviewable.

generate_chart.py reads records.json at publish time and renders a thumbnail
chip in each person's Source line.

Curation rule (Welty conventions): decoys, negative-sweep calibration pages,
and records that belong to a DIFFERENT family (the Newberry D7 Philip, the
Manheim John Welty cluster, the Donegal Michael) are marked show=False and
routed to _evidence, never silently attached to a direct ancestor.
"""
import os, re, json, subprocess, glob
from PIL import Image
import openpyxl

ROOT   = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
RECDIR = os.path.join(ROOT, "Record Images")
SITE   = os.path.join(os.path.dirname(__file__), "..", "site")
OUTF   = os.path.join(SITE, "records", "full")
OUTT   = os.path.join(SITE, "records", "thumb")
MANXL  = os.path.join(RECDIR, "Capture Manifest.xlsx")
FULL_MAX, THUMB_MAX = 1600, 360

# ---------------------------------------------------------------------------
# CURATED PER-FILE MAP  (basename -> persons / caption / confidence)
#   persons: roster PersonIDs the record is pinned to. [] => evidence-only.
#   conf:    high | med | review   (review = attribution wants Kwyn's eye)
# ---------------------------------------------------------------------------
M = {
 # --- Edenkoben Reformed registers (Archion) ------------------------------
 "P40 — Philip Jacob Wälti baptism 7 May 1719 (Archion reg 54930 Edenkoben2 Bild 71).pdf":
   (["E-philipjacob"], "Philip Jacob Welty's baptism, 7 May 1719, Edenkoben Reformed — 'Jacob Wälti & Anna Catharina', the record that proves his parentage (Bild 71).", "high"),
 "P41 — John Jacob Welde baptism 1710 (Archion reg 54930 Edenkoben2 Bild 52).pdf":
   (["E-johnjacob1710"], "John Jacob Welde's baptism, 1710, Edenkoben Reformed — parents Hanß Jacob Wäldy & Anna Catharina (Bild 52).", "high"),
 "FB19 — Georg Wolfgang Wälti baptism 29 Sep 1716 (Archion reg 54930 Edenkoben2 Bild 66).pdf":
   (["E-georgwolfgang-disp"], "Georg Wolfgang Wälti's baptism, 29 Sep 1716, Edenkoben Reformed — father Hanß Jacob Wälti & Anna Catharina (Bild 66).", "high"),
 "P53 — Johannes Wälti baptism 2 Feb 1713 (Archion reg 54930 Edenkoben2 Bild 58).pdf":
   (["E-johannes1713"], "Johannes Wälti's baptism, 2 Feb 1713, Edenkoben Reformed — father Hanß Jacob Wäldi 'Müller' & Anna Catharina (Bild 58).", "high"),
 "P53 — Anna Barbara Wälti baptism Aug 1714 (Archion reg 54930 Edenkoben2 Bild 60).pdf":
   (["E-annabarbara1714"], "Anna Barbara Wälti's baptism, Aug 1714, Edenkoben Reformed — father Hanß Jacob Wäldi & Anna Catharina (Bild 60).", "high"),
 "FB57 — Georg Wolffgang Welde marriage entry (Archion reg 54927 Edenkoben Taufen Bild 56).pdf":
   (["E-georgwolfgang-disp"], "Georg Wolfgang Welde's marriage entry, Edenkoben Reformed (Bild 56).", "high"),
 "FB57 — Georg Wolffgang Welde entry (Archion reg 54927 Edenkoben Taufen Bild 62).pdf":
   (["B-johanjacob"], "The 14 Jun 1750 Edenkoben baptism naming Georg Wolfgang Welde as father (Bild 62).", "high"),
 "P75 — Anna Elisabetha Welde baptism 31 Jan 1740 (Archion reg 54927 Edenkoben Taufen Bild 11).pdf":
   (["E-jjc-annaelis1740"], "Anna Elisabetha Welde's baptism, 31 Jan 1740, Edenkoben Reformed (Bild 11).", "high"),
 "P75 — Welde child baptism (Archion reg 54927 Edenkoben Taufen Bild 28).pdf":
   (["E-jjc-mariamarg1744"], "Maria Margretha Welde's baptism, 29 Sep 1744, Edenkoben Reformed (Bild 28).", "med"),
 "P75 — Welde child baptism 1745-46 (Archion reg 54927 Edenkoben Taufen Bild 32).pdf":
   (["E-jjc-johnicolaus1746"], "Johann Nicolaus Welde's baptism, 16 Jan 1746, Edenkoben Reformed (Bild 32).", "med"),
 "P75 — Welde child baptism (Archion reg 54927 Edenkoben Taufen Bild 43).pdf":
   (["E-jjc-annabarb1748"], "Anna Barbara Welde's baptism, 4 Feb 1748, Edenkoben Reformed (Bild 43).", "med"),
 "P74 + FB56 — Joh Jacob Welde x Anna Catharina Croissant marriage (Archion reg 54933 Edenkoben4 Bild 79).pdf":
   (["E-johnjacob1710"], "Marriage of Joh. Jacob Welde and Anna Catharina Croissant, Edenkoben Reformed (Bild 79).", "high"),
 "P64 — Edenkoben marriage register, bride's father read in full (Archion reg 54933 Edenkoben4 Bild 63).pdf":
   (["E-hansjacobwaldi","E-johanngeorgwaldi"], "Marriage of Hanß Jacob Wäldi 'der Müller', Sept 1709, Edenkoben Reformed — son of the late Joh. Georg Wäldi, former citizen of Bischheim, to Anna Catharina Durchsteinbach [P42/P64] (Bild 63). Same record: Hans Jacob's marriage and the sole life-mention of his father Johann Georg.", "high"),
 "P80 — Hanß Jacob Croissant x A.M. Müllerin marriage (Archion reg 54933 Edenkoben4 Bild 66).pdf":
   (["E-johnjacob1710"], "Marriage of Hanß Jacob Croissant & A. M. Müllerin — the Croissant in-laws (Bild 66).", "med"),
 "P116 — Doll x Neu marriage (Archion reg 54933 Edenkoben4 Bild 93).pdf":
   ([], "Doll × Neu marriage — methodology / graft-source calibration record (Bild 93).", "high"),
 "P82 — Edenkoben marriage sweep (Archion reg 54933 Edenkoben4 Bild 80).pdf":
   ([], "Calibration page for the full Edenkoben marriage-register sweep (Bild 80).", "high"),
 # --- Archive.org (free) --------------------------------------------------
 "D15 — Welty Philip 40ac Northampton 1785 decoy (Archive.org 3rdPAarch v19 p141).png":
   ([], "DECOY — a Philip Welty, 40 ac Northampton Co. 1785; ruled out, not the Dover Philip.", "high"),
 "P13 + P147 — York Co 1780 tax return p203 (Welty Widow + Philip Welty + George Welty) (archive.org 3rdPAarch v21 leaf221).jpg":
   (["E-philipjacob"], "York Co. 1780 tax return — Welty Widow, Philip Welty and George Welty (PA Arch. 3rd ser. v21, p.203).", "high"),
 "P2 — Philip Jacob Welde signature, Royal Union 1750 p433 (archive.org pennsylvaniagerm42stra leaf516).jpg":
   (["E-philipjacob"], "Philip Jacob Welty's arrival — 'Phipps Welde' on the Royal Union oath list, Philadelphia 1750 (Strassburger, Pennsylvania German Pioneers, p.433).", "high"),
 "P70 — Lydia 1749 shiplist p420 (List 142C start, Ship Lydia) (archive.org pennsylvaniagerm42stra leaf502).jpg":
   (["E-johnjacob1710"], "Ship Lydia 1749 passenger list, p.420 — the Croissant in-law crossing (List 142C).", "med"),
 "P70 — Lydia 1749 shiplist p421 (Johan Jacob Messer) (archive.org pennsylvaniagerm42stra leaf503).jpg":
   (["E-johnjacob1710"], "Ship Lydia 1749 passenger list, p.421.", "med"),
 "P70 — Lydia 1749 shiplist p422 (Jeorg Crassan = Croissant) (archive.org pennsylvaniagerm42stra leaf504).jpg":
   (["E-johnjacob1710"], "Ship Lydia 1749 list, p.422 — 'Jeorg Crassan' (Croissant), the in-law surname.", "med"),
 "US14 — Johannes Welte, Neptune 1751 shiplist p469 (archive.org pennsylvaniagerm42stra leaf553).jpg":
   ([], "UNSORTED — Johannes Welte, ship Neptune 1751; not yet tied to the line.", "high"),
 # --- FamilySearch --------------------------------------------------------
 "FA32 — Joseph Welty ~19 son of John Welty late of Manheim (FamilySearch ark 3QS7-899B-27YW).jpg":
   ([], "Separate cluster — Joseph Welty (~19), son of John Welty late of Manheim.", "high"),
 "FA33 — John Welty deceased testate - exec Christian Rubel (FamilySearch ark 3QS7-L99B-27Q1).jpg":
   ([], "Separate cluster — John Welty dec'd (testate), exec. Christian Rubel.", "high"),
 "FA34 — John Welty Sr will - wife Eve + 12-child heir list (FamilySearch ark 3QS7-899B-K845).jpg":
   ([], "Separate cluster — John Welty Sr. will, wife Eve + 12-child heir list.", "high"),
 "FB37 — Henry Welty x Coleman Manchester deed 1839 (FamilySearch York Deed Bks 1819-45 img76).png":
   (["B-henrySMGF"], "Henry Welty & Coleman, Manchester Twp deed, 1839 (York Deed Bks 1819–45, img 76).", "high"),
 "M141 — Daniel Welty household 1850 census, Marion Twp Owen Co IN (FamilySearch img291, ark S3HY-69YS-CP3).pdf":
   ([], "REVIEW — Daniel Welty household, 1850 census, Marion Twp, Owen Co. IN; which Daniel to confirm.", "review"),
 "M141 — Daniel Welty x Louisa Whaling marriage 5 Apr 1839, Tuscarawas Co OH (FamilySearch marriage reg entry 2748, ark 33S7-95BW-VCC).pdf":
   ([], "REVIEW — Daniel Welty × Louisa Whaling, 5 Apr 1839, Tuscarawas Co. OH; which Daniel to confirm.", "review"),
 "M20 — Michael Welty 1828 partition, heirs named (FamilySearch Tuscarawas Court Recs img96).png":
   (["E-michael","E-henry","E-michaeljr","E-maryj","E-david1803","E-elizabethw","E-susanna1802","E-barbara","E-daniel","E-samuel"],
    "1828 partition of Michael Welty's estate in Chancery, Tuscarawas Co. OH (Common Pleas, July term 1828) — names his children & heirs-at-law: Michael Jr. (petitioner), Mary (late Welty, wife of Peter Jacobs), David, Elizabeth (late Welty, wife of Christian Plough), Susanna, Barbara, the minors Daniel and Samuel (guardian Abraham Overholt), and Henry (son, then residing in Fayette Co., PA).", "high"),
 "M159 — Michael Welty Continental Line Depreciation Pay p481 (archive.org 5thpennsylvaniaarch04harruoft leaf486).jpg":
   (["E-michael"], "'Welty, Michael, private' on the Revolutionary War Continental Line Depreciation Pay roll (PA Archives 5th ser. Vol. IV, p.481).", "high"),
 "M28 — Michael Welty PA Rev War militia pay certificate 12706 York (Fold3 img 712667347 PHMC) p1.jpg":
   (["E-michael"], "Michael Welty's Revolutionary War militia pay certificate — 'WELTY, MICHAEL, County York, Unit Militia, Certificate 12,706, Total £10.10.0', issued 26 Oct 1786 under the Militia Loan of 1784–85 (Comptroller General records). Back-pay for his York County militia service; the Rank field is blank, refuting the family 'Lieutenant' tradition. PA Revolutionary War Military Abstract Card File (PA Historical & Museum Commission).", "high"),
 "H3 — Henry Welty & Mary Byerly family register (births marriages deaths, typed) (Welty family document IMG_1043).jpeg":
   (["E-henry","E-catherine1813","E-john1814","E-william1816","E-mary1818","E-susanna1821","E-george1823","E-nancy","E-henryjr","E-david1831"],
    "The Henry Welty & Mary Byerly family register — a typed transcript of the family's Bible-style record, held in the family. Henry Welty b. 30 Oct 1790; m. Mary Byerly (b. 3 Jun 1789) on 16 Jan 1812; the births of all nine children (Catherine 1813, John 1814, William 1816, Mary 1818, Susanna 1821, George 1823, Nancy 1826, Henry Jr. 1828, David 1831); the Dowell and Otis marriages; and the family deaths — Mary (dau.) 1822, Catherine Dowell 1840, Henry himself 8 Oct 1842, wife Mary 1848, son George 1890. The keystone record for the direct-line Gen-5 household.", "high"),
 "H17 — Henry Welty estate inventory + widows allowance, Mary Welty executrix, Stark Co OH 1842-43 (FamilySearch Probate Admin Recs 1841-43 img202, ark 3QS7-99Q2-FNZW).jpg":
   (["E-henry"], "Henry Welty's estate — the appraisement/inventory and the widow's allowance (left page). 'Personally came Mary Welty, Executrix of the last will and testament of Henry Welty, deceased' (sworn 23 Feb 1843); the appraisers set off a year's support for the widow and children (Dec 1842); an included note names son William Welty. Death/probate proof, Stark Co. Ohio Common Pleas. FamilySearch, Ohio Probate Records — Stark, Administration Records 1841–43, img 202 (ark 3QS7-99Q2-FNZW).", "high"),
 "H1 — Henry Welty baptism 1790 Center Lutheran Somerset citation (Ancestry coll2451 rec2021058430).png":
   (["E-henry"], "Baptism of Henry Welty (recorded 'Henry Wöldy'), b. 30 Oct 1790, bp. 2 Jan 1791, Center Lutheran Church, Stanton Mills, Somerset Co. PA — father Michael Wöldy, mother Christina Wöldy. The primary parentage record for the Michael → Henry generation; also pins Michael's family in Somerset Co. in 1790–91. Historic PA Church & Town Records (HSP) via Ancestry coll. 2451; citation card.", "high"),
 "H11 — Austin Welty burial 1869-1938 Six Corners Hicksville citation (FindAGrave 144975762).png":
   (["E-austin"], "Burial of Austin Welty (1869–1938), Six Corners Cemetery, Hicksville, Defiance Co. OH — son of George W. Welty & Sarah Jones, the direct-line Gen-7 rung. Wife Martha M. Kyle (m. 1905); son Robert Warren Welty (1909–1961) continues the line. Find a Grave memorial 144975762; citation card (contributor headstone photo not reproduced).", "high"),
 "H12 — Robert Warren Welty burial 1909-1961 Six Corners Hicksville citation (FindAGrave 144915461).png":
   (["E-robert"], "Burial of Robert Warren Welty (b. 24 Jan 1909 Mount Vernon, IL; d. 11 Mar 1961), Six Corners Cemetery, Hicksville, Defiance Co. OH — son of Austin Welty & Martha Kyle, the direct-line Gen-8 rung. Wife Marian Collins Sprow (1908–2001, m. 1931), bringing in the maternal Sprow line; son Dr. Alan Sprow Welty continues the line. Find a Grave memorial 144915461; citation card (contributor headstone photo not reproduced).", "high"),
 "H13 — Dr Alan Sprow Welty burial 1937-2009 Six Corners Hicksville citation (FindAGrave 144912505).png":
   (["E-alan"], "Burial of Dr. Alan Sprow Welty (b. 15 Aug 1937 Hicksville; d. 27 May 2009 Cleveland), Six Corners Cemetery, Hicksville, Defiance Co. OH — son of Robert Warren & Marian (Sprow) Welty, the direct-line Gen-9 rung. Orthodontist in Owosso MI for 30 years; Asst. Prof. of Oral Surgery, Univ. of Pittsburgh 1963–67; m. Joy Bindbeutel, 26 Mar 1960. Find a Grave memorial 144912505 + 2009 obituary; citation card.", "high"),
 "H8 — George W Welty burial 1823-1891 Forest Home Hicksville citation (FindAGrave 73762003).png":
   (["E-george1823"], "Burial of George W. Welty (b. 19 Oct 1823, d. 21 Oct 1891), Forest Home Cemetery, Hicksville, Defiance Co. OH — Henry & Mary Byerly's son and the direct-line Gen-6 rung. Wives Elizabeth Frease (m. 1852) and Sarah 'Sallie' Jones (m. 1862); children include Frank, Josephine, and Austin Welty (the line onward). This George = FTDNA Y-DNA kit #402746 (Y-67), the tested descendant confirming the Edenkoben R1b male line. Find a Grave memorial 73762003; citation card (contributor headstone photo not reproduced).", "high"),
 "H28 — Henry Welty on Connellsville Borough Town Council 1818-1827 (archive.org History of Fayette Co PA Ellis 1882 p397 leaf487).jpg":
   (["E-henry"], "Henry Welty elected to the Connellsville Borough (Fayette Co., PA) Town Council in 1818, 1819, 1824 and 1827 — the printed borough civil list. Repeated public office documents Henry's residence in Connellsville across 1818–1827, filling the Fayette-County window between his 1812 marriage and the move to Ohio. Ellis, History of Fayette County, Pennsylvania (1882), p. 397 (archive.org).", "high"),
 "H18 — Henry Welty estate partition case 224, heirs named, Stark Co OH Common Pleas 1850-51 (FamilySearch Appearance Dockets 1849-52 img262).jpg":
   (["E-henry","E-william1816","E-george1823","E-henryjr","E-david1831","E-nancy","E-susanna1821"],
    "Partition of Henry Welty's real estate among his heirs — Stark Co. Ohio Common Pleas case #224, 'In Partition': William Welty (petitioner) v. George Welty, Henry Welty, David Welty, Nancy Welty (intermarried with John Stansbury), Susan Welty (intermarried with Walter Carr) & Jeremiah Grant; petition filed 23 Aug 1850, amicably settled 7 May 1851. Names six of Henry's children as heirs. FamilySearch, Stark Co. Appearance Dockets 1849–52, img 262 (ark 3QHV-V3D4-MSSQ).", "high"),
 "M97 — Daniel Prosser to Michael Weldy Donegal deed 1797 (FamilySearch Westmoreland Deed Bks img156).png":
   ([], "REVIEW — Daniel Prosser to Michael Weldy, Donegal Twp deed 1797 (Westmoreland Co.).", "review"),
 "P110 — OC docket general index W-section Books D-K (zero Welty) (FamilySearch ark 3QSQ-G99B-VG4K).jpg":
   ([], "Negative sweep — Orphans' Court docket index, W-section, Books D–K: zero Welty.", "high"),
 "P111 — Gauff settlement motion (want of notice to guardians) (FamilySearch ark 3QS7-L99B-278P).jpg":
   (["E-elizabethgauf"], "Gauf estate settlement motion (want of notice to guardians).", "high"),
 "P111+P112 — George Gauff guardianships - Elizabeth + Margaret (FamilySearch ark 3QSQ-G99B-2W6X).jpg":
   (["E-elizabethgauf"], "George Gauf guardianships — daughters Elizabeth & Margaret.", "high"),
 "P111+P112 — OC docket general index G-section (Gauff lines) (FamilySearch ark 3QS7-L99B-VLVJ).jpg":
   (["E-elizabethgauf"], "Orphans' Court docket index, G-section — the Gauff lines.", "med"),
 "P112 — George Gauf admin account - Jacob Nigle + Catharine (FamilySearch ark 3QS7-L99B-2731).jpg":
   (["E-elizabethgauf"], "George Gauf administration account (Jacob Nigle + Catharine).", "med"),
 "P113 — Letters of administration memo - Catharine Gauff + Jacob Weigel (FamilySearch ark 3QS7-L99B-KZR3).jpg":
   (["E-elizabethgauf"], "Letters of administration — Catharine Gauff + Jacob Weigel.", "med"),
 "P114 — Jacob Lower guardianship account for Elizabeth Goff (FamilySearch ark 3QS7-L99B-2796).jpg":
   (["E-elizabethgauf"], "Jacob Lower's guardianship account for Elizabeth Goff (Gauf).", "med"),
 "P120 — QS continuance p142 Catharine bound 40 pounds (FamilySearch ark 3Q9M-CSVF-LTZL).jpg":
   (["E-elizabethgauf"], "Quarter Sessions continuance p.142 — Catharine bound £40 (Gauf matter).", "med"),
 "P120 — QS disposition case No.10 State v Catharine Gauff (FamilySearch ark 3Q9M-CSVF-LTNX).jpg":
   (["E-elizabethgauf"], "Quarter Sessions disposition, case No.10 — State v. Catharine Gauff.", "med"),
 "P120 — QS recognizances p133 vs Catharine Gauff (FamilySearch ark 3Q9M-CSVF-LTVG).jpg":
   (["E-elizabethgauf"], "Quarter Sessions recognizances p.133 — vs Catharine Gauff.", "med"),
 "P120 — QS cancelled duplicate entry (FamilySearch ark 3Q9M-CSVF-LTLK).jpg":
   ([], "Quarter Sessions cancelled duplicate entry — housekeeping image.", "high"),
 "P121 — Philip Gauf of Dover estate 868 pounds (FamilySearch ark 3QS7-L99B-2LD).jpg":
   (["E-elizabethgauf"], "Philip Gauf of Dover — estate inventory (£868); the elder Gauf household.", "med"),
 "P122 — 1774 warrant Dundohr 80ac Dover Twp (FamilySearch ark 3QHV-J385-19F7-V).jpg":
   (["E-philipjacob"], "1774 land warrant, 80 ac Dover Twp — adjoining-owner context for Philip Jacob Welty.", "med"),
 "P127 — Dover 1771 assessment continuation D-Z + Manheim warrant (FamilySearch ark 3Q9M-CSVF-Y9VN-B).jpg":
   (["E-philipjacob"], "Dover Twp 1771 assessment (D–Z) + Manheim warrant — Dover tax context.", "med"),
 "P127+P145 — Dover 1771-for-1772 assessment (verified zero-Welty + Peter warrant) (FamilySearch ark 3Q9M-CSVF-Y9VF-6).jpg":
   (["E-philipjacob"], "Dover Twp 1771-for-1772 assessment — verified zero-Welty calibration + Peter warrant.", "med"),
 "P143 — Dover 1775 Provincial Tax D-Z, Welty Philip + George (FamilySearch York tax film 008122326 img148, ark 3Q9M-CSVF-Y9VD-Y).jpg":
   (["E-philipjacob"], "Dover Twp 1775 Provincial Tax (D–Z) — Welty Philip & George.", "high"),
 "P143 — Dover 1775 Provincial Tax warrant 8 Dec 1774 + A-D (FamilySearch ark 3Q9M-CSVF-Y9VH-5).jpg":
   (["E-philipjacob"], "Dover Twp 1775 Provincial Tax — warrant of 8 Dec 1774 + A–D.", "med"),
 "P38 — Jacob Welday Sr 10 children, First Families of Ohio (FamilySearch fullText img1327).png":
   (["E-jacobsr"], "Jacob Welday Sr.'s ten children (First Families of Ohio record).", "high"),
 "P4 — Manchester Twp 1771-for-1772 assessment, W-column (FamilySearch film 008122326, ark 3Q9M-CSVF-Y9VF-J).jpg":
   (["E-georgwolfgang-disp"], "Manchester Twp 1771-for-1772 assessment — the W-column.", "high"),
 "P4 — Manchester assessment wrapper 1771-for-1772 (FamilySearch film 008122326, ark 3Q9M-CSVF-Y9VN-H).jpg":
   (["E-georgwolfgang-disp"], "Manchester Twp 1771-for-1772 assessment — cover wrapper.", "med"),
 "P4 — Manchester collectors warrant (To John Ferry Collector) (FamilySearch ark 3Q9M-CSVF-Y9VN-5).jpg":
   (["E-georgwolfgang-disp"], "Manchester Twp collector's warrant (to John Ferry, collector).", "med"),
 "P6 — Philip Welty of Newberry to Jacob Bare 1797, D7 firewall (FamilySearch York Deed Bks img200).png":
   ([], "FIREWALL — Philip Welty of Newberry Twp (the D7 Philip), a DIFFERENT man from the Dover Philip.", "high"),
 "TL55 - Welty Black Book pg3 (John Jacob 1710 x Christina Broff family register; FS memory 9VSM-X8K, uploader genehisthome).png":
   (["E-johnjacob1710"], "The 'Welty Black Book' family-register page for John Jacob Welde (b.1710). Note: the 'Broff' wife-name it carries is a debunked name-match.", "med"),
 "US19 — Obediah Weldy decoy, Chester tax 1768-71 (FamilySearch img109).png":
   ([], "DECOY — Obediah Weldy, Chester Co. tax 1768–71; not the line.", "high"),
 # --- Session 12 Jul 2026: Gen 1-4 photo pass (BMD keys) -------------------
 "M18 — Michael Weldy estate inventory, deceased Sept 1815 (FamilySearch Tuscarawas Admin Recs 1810-39 film 005870723 img19, ark 3QS7-99Q2-XJTK).jpg":
   (["E-michael"], "Estate inventory of Michael Weldy, deceased Sept 1815, Tuscarawas Co. OH — the appraisement of his goods and property; a death-bracketing record (FamilySearch, Administration Records 1810–39, img 19).", "high"),
 "FB20 — Georg Wolfgang Weldy x Anna Maria Keibel marriage 7 Jan 1750 (Archion reg 54933 Edenkoben4 Bild 89).pdf":
   (["E-georgwolfgang-disp"], "Marriage of Georg Wolfgang Weldy & Anna Maria Keibel of Deggelsheim, 7 Jan 1750, Edenkoben Reformed — the primary marriage record, entered under the '1750' year heading (Bild 89).", "high"),
 "M4 — Michael Welty x Christina Ruthrauff marriage 18 May 1784 citation (Ancestry coll4946 ev7965).png":
   (["E-michael"], "Marriage of Michael Welty & Christina Ruthrauff (recorded 'Christine Rohtrauf'), 18 May 1784, First Reformed (Trinity) Church, York — the corrected date superseding the pedigree's '10 May'. Register transcription (Ancestry coll. 4946, event 7965); citation card.", "high"),
 "P160 — John Henry Welty burial citation, d 4 Jun 1841 (FindAGrave 32007221 St Clair Cem Greensburg).png":
   (["E-johnhenry1764"], "Burial of Johan Heinrich 'Henry' Welty Sr. (b. 4 Nov 1764, d. 4 Jun 1841), Saint Clair Cemetery, Greensburg, Westmoreland Co. PA; wife Eva Catherine Steiner. Find a Grave memorial 32007221; citation card (contributor headstone photo not reproduced).", "med"),
 "P126 — Philip Welty & wife Elisabet sponsor 1786 citation (FS-DL 892477 img29 Strayers).png":
   (["E-philipjacob"], "Philip Welty & his wife Elisabet stand sponsors at Daniel Glaser's baptism, 9 Nov 1786, Strayer's (Salem) Reformed, Dover Twp — the only primary record naming Philip's wife Elisabet. Young/Byrnes transcription (FS-DL 892477, img 29); citation card.", "high"),
 # --- York daughters' marriages, Trinity York (Ancestry coll.4946) ---------
 "M68 — Elizabeth Welty x George Gauf marriage 26 Mar 1775 citation (Ancestry coll4946 ev7751).png":
   (["E-elizabethgauf"], "Marriage of Elizabeth Welty & George Gauf, 26 Mar 1775, First Reformed (Trinity) Church, York — recorded 'Welty Gauf'. Register transcription (Ancestry coll. 4946, event 7751); citation card.", "high"),
 "M72 — Christina Welty x Peter Messerle marriage 13 Apr 1784 citation (Ancestry coll4946 ev7962).png":
   (["E-christinamesserle"], "Marriage of Christina Welty & Peter Messerle, 13 Apr 1784, First Reformed (Trinity) Church, York. Register transcription (Ancestry coll. 4946, event 7962); citation card.", "high"),
 "US16 — Catharine Welty x Jacob Boehm marriage 10 May 1785 citation (Ancestry coll4946 ev7987).png":
   (["E-catharinaboehm"], "Marriage of Catharine Welty & Jacob Boehm, 10 May 1785, First Reformed (Trinity) Church, York — part of the 10 May 1785 double Welty wedding (the likely origin of the ghost '10 May' date once mis-copied onto Michael's 1784 marriage). Register transcription (Ancestry coll. 4946, event 7987); citation card.", "high"),
 # --- Edenkoben confirmation lists (Archion reg 54930) ---------------------
 "P54 — Hans Philipp Waldi confirmation Ostern 1707 (Archion reg 54930 Edenkoben2 Bild 127).pdf":
   (["E-hansphilipp1693","E-johanngeorgwaldi"], "Confirmation of Hanß Philipp Wäldi, Easter 1707, Edenkoben Reformed — entered as 'Hanß Philipps Wäldin, 14, des weiland Hanß Görg Wäldins zu Bischheim nachgel. Sohn': a second independent record naming the late Johann Georg Wäldi of Bischheim as father (Bild 127).", "high"),
 "P55 — Johann Jacob Waldy confirmation Ostern 1726 (Archion reg 54930 Edenkoben2 Bild 131).pdf":
   (["E-johnjacob1710"], "Confirmation of Johann Jacob Wäldy, Easter 1726, Edenkoben Reformed — the 6th boy in the Anno-1726 Knaben list; same spelling as his 1710 baptism, a second Edenkoben primary record for him (Bild 131).", "high"),
}

# ---------------------------------------------------------------------------
# VITAL-ONLY POLICY (Kwyn, 11 Jul 2026): a record earns a chip on a person only
# if it's a vital event FOR THEM (baptism/birth, marriage, death/burial, will,
# partition/guardianship) or the keystone record that NAMES them where little
# else does. Everything else — in-laws, elder-Gauf & Catharine court records,
# adjoining-owner warrants, negative/calibration pages, and second views of a
# fact already shown — stays in the Record Images backlog as evidence-only.
# To promote a record onto a person, add its filename here.
KEEP_ON_TREE = {
 "P40 — Philip Jacob Wälti baptism 7 May 1719 (Archion reg 54930 Edenkoben2 Bild 71).pdf",                                # baptism — the parentage proof
 "P41 — John Jacob Welde baptism 1710 (Archion reg 54930 Edenkoben2 Bild 52).pdf",                                        # baptism
 "FB19 — Georg Wolfgang Wälti baptism 29 Sep 1716 (Archion reg 54930 Edenkoben2 Bild 66).pdf",                            # baptism
 "P53 — Johannes Wälti baptism 2 Feb 1713 (Archion reg 54930 Edenkoben2 Bild 58).pdf",                                    # baptism
 "P53 — Anna Barbara Wälti baptism Aug 1714 (Archion reg 54930 Edenkoben2 Bild 60).pdf",                                  # baptism
 "TL55 - Welty Black Book pg3 (John Jacob 1710 x Christina Broff family register; FS memory 9VSM-X8K, uploader genehisthome).png",  # family register in his own hand
 "P74 + FB56 — Joh Jacob Welde x Anna Catharina Croissant marriage (Archion reg 54933 Edenkoben4 Bild 79).pdf",           # marriage
 "FB57 — Georg Wolffgang Welde marriage entry (Archion reg 54927 Edenkoben Taufen Bild 56).pdf",                          # marriage
 "FB57 — Georg Wolffgang Welde entry (Archion reg 54927 Edenkoben Taufen Bild 62).pdf",                                   # 1750 baptism of Johan Jacob
 "P75 — Anna Elisabetha Welde baptism 31 Jan 1740 (Archion reg 54927 Edenkoben Taufen Bild 11).pdf",                      # baptism
 "M20 — Michael Welty 1828 partition, heirs named (FamilySearch Tuscarawas Court Recs img96).png",                        # partition
 "M159 — Michael Welty Continental Line Depreciation Pay p481 (archive.org 5thpennsylvaniaarch04harruoft leaf486).jpg",   # Rev War depreciation pay
 "M28 — Michael Welty PA Rev War militia pay certificate 12706 York (Fold3 img 712667347 PHMC) p1.jpg",                   # Rev War militia pay certificate (PHMC abstract card)
 "H3 — Henry Welty & Mary Byerly family register (births marriages deaths, typed) (Welty family document IMG_1043).jpeg",  # Gen-5 keystone: marriage + 9 kids + deaths (family document)
 "H17 — Henry Welty estate inventory + widows allowance, Mary Welty executrix, Stark Co OH 1842-43 (FamilySearch Probate Admin Recs 1841-43 img202, ark 3QS7-99Q2-FNZW).jpg",  # Gen-5 death/probate: estate inventory + widow's allowance
 "H18 — Henry Welty estate partition case 224, heirs named, Stark Co OH Common Pleas 1850-51 (FamilySearch Appearance Dockets 1849-52 img262).jpg",  # Gen-5 heirs: partition names 6 children
 "H1 — Henry Welty baptism 1790 Center Lutheran Somerset citation (Ancestry coll2451 rec2021058430).png",  # Gen-5 birth/parentage (citation card)
 "H28 — Henry Welty on Connellsville Borough Town Council 1818-1827 (archive.org History of Fayette Co PA Ellis 1882 p397 leaf487).jpg",  # Gen-5 residence: Connellsville town council 1818-27
 "H8 — George W Welty burial 1823-1891 Forest Home Hicksville citation (FindAGrave 73762003).png",  # Gen-6 death/burial (citation card) + Y-DNA anchor
 "H11 — Austin Welty burial 1869-1938 Six Corners Hicksville citation (FindAGrave 144975762).png",  # Gen-7 death/burial (citation card)
 "H12 — Robert Warren Welty burial 1909-1961 Six Corners Hicksville citation (FindAGrave 144915461).png",  # Gen-8 death/burial (citation card)
 "H13 — Dr Alan Sprow Welty burial 1937-2009 Six Corners Hicksville citation (FindAGrave 144912505).png",  # Gen-9 death/burial (citation card)
 "P111+P112 — George Gauff guardianships - Elizabeth + Margaret (FamilySearch ark 3QSQ-G99B-2W6X).jpg",                   # guardianship names her line
 "P38 — Jacob Welday Sr 10 children, First Families of Ohio (FamilySearch fullText img1327).png",                         # names his children
 "P13 + P147 — York Co 1780 tax return p203 (Welty Widow + Philip Welty + George Welty) (archive.org 3rdPAarch v21 leaf221).jpg",  # names Philip; resolves the 1780 Widow
 "P2 — Philip Jacob Welde signature, Royal Union 1750 p433 (archive.org pennsylvaniagerm42stra leaf516).jpg",                    # immigration oath — Phipps Welde 1750
 "FB37 — Henry Welty x Coleman Manchester deed 1839 (FamilySearch York Deed Bks 1819-45 img76).png",                      # keystone deed naming this Henry
 "P75 — Welde child baptism (Archion reg 54927 Edenkoben Taufen Bild 28).pdf",                                            # baptism — Maria Margretha 1744
 "P75 — Welde child baptism 1745-46 (Archion reg 54927 Edenkoben Taufen Bild 32).pdf",                                   # baptism — Johann Nicolaus 1746
 "P75 — Welde child baptism (Archion reg 54927 Edenkoben Taufen Bild 43).pdf",                                           # baptism — Anna Barbara 1748
 "P64 — Edenkoben marriage register, bride's father read in full (Archion reg 54933 Edenkoben4 Bild 63).pdf",            # Hans Jacob's 1709 marriage + sole life-mention of father Joh. Georg
 "M18 — Michael Weldy estate inventory, deceased Sept 1815 (FamilySearch Tuscarawas Admin Recs 1810-39 film 005870723 img19, ark 3QS7-99Q2-XJTK).jpg",  # death-bracket: estate inventory
 "FB20 — Georg Wolfgang Weldy x Anna Maria Keibel marriage 7 Jan 1750 (Archion reg 54933 Edenkoben4 Bild 89).pdf",       # marriage
 "M4 — Michael Welty x Christina Ruthrauff marriage 18 May 1784 citation (Ancestry coll4946 ev7965).png",                # marriage (citation card)
 "P160 — John Henry Welty burial citation, d 4 Jun 1841 (FindAGrave 32007221 St Clair Cem Greensburg).png",              # burial (citation card)
 "P126 — Philip Welty & wife Elisabet sponsor 1786 citation (FS-DL 892477 img29 Strayers).png",                         # names wife Elisabet (citation card)
 "M68 — Elizabeth Welty x George Gauf marriage 26 Mar 1775 citation (Ancestry coll4946 ev7751).png",                     # daughter marriage (citation card)
 "M72 — Christina Welty x Peter Messerle marriage 13 Apr 1784 citation (Ancestry coll4946 ev7962).png",                  # daughter marriage (citation card)
 "US16 — Catharine Welty x Jacob Boehm marriage 10 May 1785 citation (Ancestry coll4946 ev7987).png",                    # daughter marriage (citation card)
 "P54 — Hans Philipp Waldi confirmation Ostern 1707 (Archion reg 54930 Edenkoben2 Bild 127).pdf",                        # confirmation (names father Johann Georg)
 "P55 — Johann Jacob Waldy confirmation Ostern 1726 (Archion reg 54930 Edenkoben2 Bild 131).pdf",                        # confirmation
}

def slugify(base):
    stem = os.path.splitext(base)[0]
    stem = stem.replace("+"," ").replace("×","x")
    stem = re.sub(r"\(.*?\)", "", stem)          # drop parenthetical repo refs
    stem = re.sub(r"[^A-Za-z0-9]+","-", stem).strip("-").lower()
    return stem[:70].strip("-")

def unique_slugs(files):
    """Deterministic slugs; siblings that collapse to the same base get -2/-3.
    First (sorted) file keeps the clean base name, so re-runs overwrite in place."""
    out={}; seen={}
    for f in files:
        b=slugify(os.path.basename(f))
        seen[b]=seen.get(b,0)+1
        out[f]= b if seen[b]==1 else f"{b}-{seen[b]}"
    return out

def load_manifest_urls():
    urls={}
    if not os.path.exists(MANXL): return urls
    wb=openpyxl.load_workbook(MANXL, read_only=True, data_only=True)
    ws=wb["Capture Manifest"]; rows=list(ws.iter_rows(values_only=True))
    h={c:i for i,c in enumerate(rows[0])}
    for r in rows[1:]:
        lid=r[h['Log ID']]
        if lid: urls[str(lid).strip()]=(r[h['Repository']], r[h['URL']])
    return urls

def to_jpg(src, slug):
    """Return a PIL image (first page for PDFs)."""
    if src.lower().endswith(".pdf"):
        import tempfile
        tmp=os.path.join(tempfile.gettempdir(), "welty_rec_tmp")
        subprocess.run(["pdftoppm","-jpeg","-r","150","-f","1","-l","1",src,tmp],
                       check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        cand=sorted(glob.glob(tmp+"*"))
        img=Image.open(cand[0]).convert("RGB")
        loaded=img.copy()
        img.close()
        for c in cand:
            try: os.remove(c)
            except OSError: pass
        return loaded
    return Image.open(src).convert("RGB")

def resized(img, mx):
    w,h=img.size
    if max(w,h)>mx:
        s=mx/max(w,h); img=img.resize((int(w*s),int(h*s)), Image.LANCZOS)
    return img

def main():
    os.makedirs(OUTF, exist_ok=True); os.makedirs(OUTT, exist_ok=True)
    urls=load_manifest_urls()
    files=sorted(f for f in glob.glob(os.path.join(RECDIR,"**","*"), recursive=True)
                 if os.path.isfile(f) and f.lower().endswith((".jpg",".jpeg",".png",".pdf")))
    slugmap=unique_slugs([f for f in files if os.path.basename(f) in M])
    by_person={}; flat_by_person={}; evidence=[]; rows_out=[]; unmapped=[]
    for f in files:
        base=os.path.basename(f)
        if base not in M:
            unmapped.append(base); continue
        persons, caption, conf = M[base]
        if base not in KEEP_ON_TREE:      # vital-only policy: demote to evidence
            persons = []
        logids=re.findall(r'(?:P|M|D|FB|FA|US)\d+|TL-?\d+', base)
        slug=slugmap[f]
        img=to_jpg(f, slug)
        resized(img, FULL_MAX).save(os.path.join(OUTF,slug+".jpg"),"JPEG",quality=82)
        resized(img, THUMB_MAX).save(os.path.join(OUTT,slug+".jpg"),"JPEG",quality=80)
        repo,url = urls.get(logids[0], (None,None)) if logids else (None,None)
        plid = logids[0] if logids else slug   # grouping key: pages of one document share a log-ID
        rec={"slug":slug,"caption":caption,"logids":logids,"repo":repo,"url":url,
             "confidence":conf,"_plid":plid}
        rows_out.append((base, ";".join(persons), ";".join(logids), caption, "yes" if persons else "no", conf, repo or "", url or ""))
        if persons:
            for p in persons: flat_by_person.setdefault(p,[]).append(rec)
        else:
            evidence.append({k:v for k,v in rec.items() if k!="_plid"} | {"file":base})

    # Collapse multiple pages of ONE document (same primary log-ID) into a single
    # chip per person, so the tree shows one thumbnail per record instead of a row
    # of near-identical manuscript pages. The extra pages ride along in `pages[]`
    # and the lightbox pages through them.
    for pid, recs in flat_by_person.items():
        groups={}
        for r in recs:
            groups.setdefault(r["_plid"], []).append(r)
        out_list=[]
        for k, items in groups.items():
            items=sorted(items, key=lambda r:r["slug"])
            head=items[0]
            out_list.append({
                "slug":head["slug"], "caption":head["caption"],
                "repo":head["repo"], "url":head["url"],
                "confidence":head["confidence"], "logids":head["logids"],
                "pages":[{"slug":r["slug"],"caption":r["caption"]} for r in items],
            })
        by_person[pid]=out_list
    chips=sum(len(v) for v in by_person.values())
    imgs_on_people=sum(len(g["pages"]) for v in by_person.values() for g in v)
    out={"by_person":by_person,"_evidence":evidence,
         "_meta":{"total_files":len(files),"chips":chips,"images_on_people":imgs_on_people,
                  "evidence_only":len(evidence),"unmapped":unmapped}}
    with open(os.path.join(SITE,"records","records.json"),"w",encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=1)
    # mirror into Capture Manifest workbook (durable, reviewable)
    wb=openpyxl.load_workbook(MANXL)
    if "Tree Image Map" in wb.sheetnames: del wb["Tree Image Map"]
    ws=wb.create_sheet("Tree Image Map")
    ws.append(["File","PersonIDs","LogIDs","Caption","Show on tree","Confidence","Repository","URL"])
    for row in sorted(rows_out): ws.append(list(row))
    wb.save(MANXL)
    print(f"files={len(files)} mapped={len(rows_out)} unmapped={unmapped}")
    print(f"chips={chips} (from {imgs_on_people} images) on {len(by_person)} people; evidence_only={len(evidence)}")

if __name__=="__main__":
    main()
