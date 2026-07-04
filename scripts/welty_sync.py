#!/usr/bin/env python3
"""SYNC step of the Welty "little git" workflow.

Google Sheet "Welty People Roster (chart source)" is the LIVE master that Kwyn &
Aaron edit. This script takes a freshly-downloaded copy of that sheet (as CSV),
writes it into the local research log's People Roster tab, and rebuilds both charts.

Usage:
    python3 welty_sync.py "People Roster - from Drive.csv"

The download-from-Drive and upload-of-charts-to-Drive steps are done by Claude via
the Google Drive connector; this script is the deterministic middle: CSV -> xlsx tab
-> regenerated HTML. Robust to column re-ordering (maps by header name).
"""
import sys, os, shutil, csv, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import generate_chart  # lives in the same folder

XLSX  = "Welty Ancestry Research Log.xlsx"
SHEET = "People Roster (chart source)"
COLS  = ["PersonID","Family","Gen","Name","Sex","Birth","Death","Place",
         "Spouse","FatherID","Proof","DNAkit","Direct","Notes"]

def rebuild_roster_tab(csv_path):
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        raise SystemExit(f"{csv_path} is empty")
    header = [h.strip() for h in rows[0]]
    missing = [c for c in COLS if c not in header]
    if missing:
        raise SystemExit(f"CSV is missing required columns: {missing}\nGot header: {header}")
    idx = {c: header.index(c) for c in COLS}
    people = []
    for r in rows[1:]:
        if not any(cell.strip() for cell in r):
            continue
        rec = [(r[idx[c]].strip() if idx[c] < len(r) else "") for c in COLS]
        if not rec[0]:   # no PersonID -> skip
            continue
        people.append(rec)

    # timestamped backup so every sync is recoverable (your commit history)
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M")
    shutil.copy2(XLSX, f"Welty Ancestry Research Log.SYNC-{stamp}.xlsx")

    wb = openpyxl.load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, index=2)

    ws["A1"] = ("People Roster — LIVE MASTER is the Google Sheet 'Welty People Roster (chart source)'. "
                "This tab is a synced copy; edit the Google Sheet, then re-sync. "
                "Add a person = add a row (fill FatherID!). Proof: proven/documented/hypo/lore/living. "
                "Direct: 'yes'=our direct line, 'kit'=Y-DNA kit owner.")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    ws["A1"].font = Font(bold=True, size=11, color="5A4632")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 46

    thin = Side(style="thin", color="C9BFAE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, c in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="4A4238")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    fam_fill = {"Eden":"FFF1EF","Manch":"EEF2F7","Swiss":"F5EEF5"}
    for i, rec in enumerate(people, start=3):
        for j, val in enumerate(rec, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            cell.font = Font(size=10)
            cell.fill = PatternFill("solid", fgColor=fam_fill.get(rec[1], "FFFFFF"))
        if rec[12] in ("yes", "kit"):
            ws.cell(row=i, column=4).font = Font(size=10, bold=True, color="B71C1C")

    widths = {"A":14,"B":8,"C":6,"D":26,"E":6,"F":13,"G":13,"H":30,"I":26,"J":14,"K":11,"L":18,"M":8,"N":60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    wb.save(XLSX)
    return len(people)

def main():
    csv_path = sys.argv[1] if len(sys.argv) > 1 else "People Roster - from Drive.csv"
    if not os.path.exists(csv_path):
        raise SystemExit(f"Cannot find {csv_path}. Claude downloads the Google Sheet to this path first.")
    n = rebuild_roster_tab(csv_path)
    generate_chart.main()
    print(f"\nSYNCED {n} people: {csv_path} -> roster tab -> both charts rebuilt.")

if __name__ == "__main__":
    main()
